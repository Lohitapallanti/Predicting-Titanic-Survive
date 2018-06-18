from openpyxl import load_workbook
from openpyxl import Workbook
import csv

'''This file analyses the input file based on learnt records, learned during machine learning part exeution'''

class Processing():

    def __init__(self):

        self.wb = load_workbook('./datasets/savedProcessed/trainProcessed.xlsm')
        self.ws = self.wb['Record_of_training']
        self.inf_male_class = [0, 0, 0]
        self.inf_female_class = [0, 0, 0]
        self.inf_cabin = [0, 0, 0, 0, 0, 0]
        self.inf_embarked = [0, 0, 0]
        for i in range(0, 3):
            self.inf_male_class[i] = self.ws.cell(row=i+2,column=1).value
            self.inf_female_class[i] = self.ws.cell(row=i + 2, column=2).value
            self.inf_embarked[i] = self.ws.cell(row = i+2, column=4).value
        for i in range(0,6):
            self.inf_cabin[i] = self.ws.cell(row = i+2, column = 3).value


    def input_file_storage(self):

        self.wb = load_workbook('./datasets/savedProcessed/trainProcessed.xlsm')
        self.ws = self.wb['Record_of_training']
        self.inf_male_class = [0, 0, 0]
        self.inf_female_class = [0, 0, 0]
        self.inf_cabin = [0, 0, 0, 0, 0, 0]
        self.inf_embarked = [0, 0, 0]
        for i in range(0, 3):
            self.inf_male_class[i] = self.ws.cell(row=i + 2, column=1).value
            self.inf_female_class[i] = self.ws.cell(row=i + 2, column=2).value
            self.inf_embarked[i] = self.ws.cell(row=i + 2, column=4).value
        for i in range(0, 6):
            self.inf_cabin[i] = self.ws.cell(row=i + 2, column=3).value

        self.parameters = []

        fname = input('.csv File name to be predicted :')
        self.fname2= fname
        ofile = open('./datasets/'+fname+'.csv', 'r')
        self.reader = csv.reader(ofile)
        self.name = []
        self.id = []
        self.pclass = []  
        self.sex = []
        self.name = []
        self.embark = []
        self.cabin = []
        self.age = []
        self.age = []
        self.SibSp = []
        self.parch = []
        self.ticket = []
        self.fare = []
        self.parameters = []


        self.survival = []  
        self.cal_sur = [] # calculating survival
        ''' use only this while checking the probability with the output and getting correct percent of probability
        self.survived = []
        for i in self.reader:
            self.id.append(i[0])

            self.survived.append(i[1])

            self.pclass.append(i[2])
            self.name.append(i[3])
            self.sex.append(i[4])
            self.age.append(i[5])
            self.cabin.append(i[10])
            self.embark.append(i[11])
            self.parameters.append(0)
            self.cal_sur.append(0)
            self.survival.append(0)
        self.count = len(self.id)
        
        '''
        for i in self.reader:
            self.id.append(i[0])

            self.pclass.append(i[1])
            self.name.append(i[2])
            self.sex.append(i[3])
            self.age.append(i[4])
            self.SibSp.append(i[5])
            self.parch.append(i[6])
            self.ticket.append(i[7])
            self.fare.append(i[8])
            self.cabin.append(i[9])
            self.embark.append(i[10])
            self.parameters.append(0)
            self.cal_sur.append(0)
            self.survival.append(0)
        self.count = len(self.id)
        self.analysis()

    def analysis(self):

        for i in range(0, self.count):

            if self.sex[i] == 'male':
                if self.pclass[i] == '1':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[0] + self.cal_sur[i]

                if self.pclass[i] == '2':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[1]+ self.cal_sur[i]

                if self.pclass[i] == '3':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[2]+ self.cal_sur[i]

            if self.sex[i] == 'female':
                if self.pclass[i] == '1':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[0]+ self.cal_sur[i]

                if self.pclass[i] == '2':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[1]+ self.cal_sur[i]

                if self.pclass[i] == '3':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[2]+ self.cal_sur[i]
            if self.cabin[i] == '':
                a=0
            else:
                if self.cabin[i][0:1] == 'A':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[0] + self.cal_sur[i]

                if self.cabin[i][0:1] == 'B':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[1] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'C':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[2] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'D':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[3] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'E':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[4] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'F':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[5] + self.cal_sur[i]

            if self.embark[i] == 'S':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[0]
            if self.embark[i] == 'C':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[1]
            if self.embark[i] == 'Q':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[2]

        # end of loop i here
        self.final_outcome()

    def final_outcome(self):
        self.probability_survival = []

        for i in range(0, self.count-1):
            #print(self.parameters[i])
            self.probability_survival.append(self.cal_sur[i] / self.parameters[i])
        no =len(self.probability_survival)
        for i in range(0, self.count-1):
            if self.probability_survival[i] >= 0.41 :
                ''' 0.41 probability value is achieved when the probability result is compared the original results of 
                training using train.py, since all of almost maximum of values above 0.41 survived 
                in the accident, hence 0.41 is taken as the bar value. '''
                self.survival[i] = 1
            else :
                self.survival[i] = 0
            if self.age[i+1] >= str(50):
                self.survival[i] = 0
        self.exporting_to_excel_file()


    def exporting_to_excel_file(self):
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = 'Predicted_Survived'
        ws2.cell(row=1, column=1,value='Id')
        ws2.cell(row=1, column=2, value='Predicted_Survival')
        ws2.cell(row=1, column=3, value='Pclass')
        ws2.cell(row=1, column=4, value='Name')
        ws2.cell(row=1, column=5, value='Sex')
        ws2.cell(row=1, column=6, value='Age')
        ws2.cell(row=1, column=7, value='SibSp')
        ws2.cell(row=1, column=8, value='Parch')
        ws2.cell(row=1, column=9, value='Ticket')
        ws2.cell(row=1, column=10, value='Fare')
        ws2.cell(row=1, column=11, value='Cabin')
        for i in range(1, self.count):
            ws2.cell(row=i+1, column=1, value=self.id[i])
            ws2.cell(row=i+1, column=2, value=self.survival[i-1])
            ws2.cell(row=i+1, column=3, value=self.pclass[i])
            ws2.cell(row=i+1, column=4, value=self.name[i])
            ws2.cell(row=i+1, column=5, value=self.sex[i])
            ws2.cell(row=i+1, column=6, value=self.age[i])
            ws2.cell(row=i+1, column=7, value=self.SibSp[i])
            ws2.cell(row=i+1, column=8, value=self.parch[i])
            ws2.cell(row=i+1, column=9, value=self.ticket[i])
            ws2.cell(row=i+1, column=10, value=self.fare[i])
            ws2.cell(row=i+1, column=11, value=self.cabin[i])
            #print(str(self.id[i]) + ' ' +str(self.survival[i-1]) + ' ' + str(self.probability_survival[i-1]))
        wb2.save('./datasets/predicted/predicted_'+self.fname2+'.xlsm')
        print('File successfully created with address ./datasets/predicted/predicted_'+self.fname2+'.xlsm')




