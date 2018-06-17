import csv
import random
from openpyxl import Workbook 
'''csv_file = open('./datasets/train.csv', 'r')
reader = csv.reader(csv_file)
for row in reader:
    print(row)'''

class train:
    def __init__(self):
        self.wb = Workbook()  
        self.ws = self.wb.active 
        self.ws.title = "Record_of_training" 



        self.survival_final = 0 
        self.id = []  
        
        
        self.survived = []
        self.pclass = []
        self.name = [] 
        self.sex = []
        self.age = []
        self.sibsp = [] 
        self.parch = [] 
        self.ticket = [] 
        self.fare = []
        self.cabin = [] 
        self.embarked = [] 

        self.male_class = [0] 
        self.male_class_influence = [0,0,0] 
        self.male_class_count = [0] * 3 
        self.female_class = [0] * 3  
        self.female_class_influence = [0,0,0] 
        self.female_class_count = [0] * 3 
        self.female_class = [0,0,0] 
        self.male_class = [0,0,0] 
        self.male_class_count = [0,0,0] 
        self.female_class_count = [0, 0, 0]


        
        self.cabin_count = 0 
        self.cabin_count_alpha = [6] 
        self.cabin_alpha_influence = [6]
        self.cabin_count_alpha_survived = [6]

        
        self.embarked_place_count = [0] * 3 
        self.embarked_place_count_survived = [0] * 3 
        self.embarked_influence = [3]

    def asking_values(self):

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Record_of_training"

        self.survival_final = 0  

        self.id = []
        self.survived = []
        self.pclass = []
        self.name = []
        self.sex = []
        self.age = []
        self.sibsp = []
        self.parch = []
        self.ticket = []
        self.fare = []
        self.cabin = []
        self.embarked = []

       
        self.cabin_count = 0
        self.cabin_count_alpha = [6]  
        self.cabin_alpha_influence = [6]
        self.cabin_count_alpha_survived = [6]

       
        self.embarked_place_count = [3]  
        self.embarked_place_count_survived = [3]
        self.embarked_influence = [3]

        self.fname = input('Enter .csv filename to train the machine : ')
        ifile = open('./datasets/'+self.fname+'.csv', 'r')
        self.reader = csv.reader(ifile) 
        '''self.csv_file = open('./datasets/train.csv', 'r')
        self.reader = csv.reader(self.csv_file)'''

        for i in self.reader:
            self.id.append(str(i[0]))
            self.survived.append(i[1])
            self.pclass.append(i[2])
            self.name.append(i[3])
            self.sex.append(i[4])
            self.age.append(i[5])
            self.sibsp.append(i[6])
            self.parch.append(i[7])
            self.ticket.append(i[8])
            self.fare.append(i[9])
            self.cabin.append(i[10])
            self.embarked.append(i[11])
        self.count = len(self.id) 
        ifile.close()


    def assigningValues(self):

        '''for i in range(1, len(self.id)):
            print(self.id[i] + ' ' + self.survived[i] + ' ' + self.pclass[i] + ' ' + self.name[i])'''
        self.calculating_influence()

    def calculating_influence(self):

       

        self.embarked_influence = [0,0,0]
        self.embarked_place_count = [0,0,0]
        self.embarked_place_count_survived = [0,0,0]
        self.cabin_alpha_influence = [0,0,0,0,0,0]
        self.cabin_count_alpha = [0,0,0,0,0,0]
        self.cabin_count_alpha_survived = [0,0,0,0,0,0]
        self.cabin_used = [False, False, False, False, False, False] 



        for i in range(1, self.count):
            if self.sex[i]=='male' :

                

                if self.pclass[i] == '1':
                     if self.survived[i] == '1' :
                        self.male_class[0] = self.male_class[0] + 1
                     self.male_class_count[0] = self.male_class_count[0] + 1

                if self.pclass[i] == '2':
                     if self.survived[i] == '1':
                        self.male_class[1] = self.male_class[1] + 1
                     self.male_class_count[1] = self.male_class_count[1] + 1

                if self.pclass[i] == '3':
                     if self.survived[i] == '1':
                        self.male_class[2] = self.male_class[2] + 1
                     self.male_class_count[2] = self.male_class_count[2] + 1

                
                ''' write this code after the loop of i ends
                self.male_class_influence[0] = self.male_class[0] / self.male_class_count[0] # class 1
                self.male_class_influence[1] = self.male_class[1] / self.male_class_count[1] # class 2
                self.male_class_influence[2] = self.male_class[2] / self.male_class_count[2] # class 3
                '''

            if self.sex[i]=='female' :

                

                if self.pclass[i] == '1':
                     if self.survived[i] == '1' :
                        self.female_class[0] = self.female_class[0] + 1
                     self.female_class_count[0] = self.female_class_count[0] + 1

                if self.pclass[i] == '2':
                     if self.survived[i] == '1':
                        self.female_class[1] = self.female_class[1] + 1
                     self.female_class_count[1] = self.female_class_count[1] + 1

                if self.pclass[i] == '3':
                     if self.survived[i] == '1':
                        self.female_class[2] = self.female_class[2] + 1
                     self.female_class_count[2] = self.female_class_count[2] + 1

               
                '''
                self.female_class_influence[0] = self.female_class[0] / self.female_class_count[0] # class 1
                self.female_class_influence[1] = self.female_class[1] / self.female_class_count[1] # class 2
                self.female_class_influence[2] = self.female_class[2] / self.female_class_count[2] # class 3
                '''
                

            
            if self.cabin == '':
                a = 0 
            else:
                self.cabin_count = self.cabin_count + 1

                if self.cabin[i][0:1] == 'A' :
                    self.cabin_count_alpha[0] = self.cabin_count_alpha[0] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[0] = self.cabin_count_alpha_survived[0] + 1
                    self.cabin_used[0] = True

                if self.cabin[i][0:1] == 'B' :

                    self.cabin_count_alpha[1] = self.cabin_count_alpha[1] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[1] = self.cabin_count_alpha_survived[1] + 1
                    self.cabin_used[1] = True

                if self.cabin[i][0:1] == 'C' :
                    self.cabin_count_alpha[2] = self.cabin_count_alpha[2] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[2] = self.cabin_count_alpha_survived[2] + 1
                    self.cabin_used[2] = True

                if self.cabin[i][0:1] == 'D' :
                    self.cabin_count_alpha[3] = self.cabin_count_alpha[3] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[3] = self.cabin_count_alpha_survived[3] + 1
                    self.cabin_used[3] = True

                if self.cabin[i][0:1] == 'E' :
                    self.cabin_count_alpha[4] = self.cabin_count_alpha[4] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[4] = self.cabin_count_alpha_survived[4] + 1
                    self.cabin_used[4] = True

                if self.cabin[i][0:1] == 'F' :
                    self.cabin_count_alpha[5] = self.cabin_count_alpha[5] + 1
                    if self.survived[i] == '1':
                        self.cabin_count_alpha_survived[5] = self.cabin_count_alpha_survived[5] + 1
                    self.cabin_used[5] = True

                '''write this code after the loop of i ends
                self.cabin_alpha_influence[0] = self.cabin_count_alpha_survived[0] / self.cabin_count_alpha[0]
                self.cabin_alpha_influence[1] = self.cabin_count_alpha_survived[1] / self.cabin_count_alpha[1]
                self.cabin_alpha_influence[2] = self.cabin_count_alpha_survived[2] / self.cabin_count_alpha[2]
                self.cabin_alpha_influence[3] = self.cabin_count_alpha_survived[3] / self.cabin_count_alpha[3]
                self.cabin_alpha_influence[4] = self.cabin_count_alpha_survived[4] / self.cabin_count_alpha[4]
                self.cabin_alpha_influence[5] = self.cabin_count_alpha_survived[5] / self.cabin_count_alpha[5]
                '''

           

            if self.embarked[i] == 'S' :
                if self.survived[i] == '1' :
                    self.embarked_place_count_survived[0] = self.embarked_place_count_survived[0] +1
                self.embarked_place_count[0] = self.embarked_place_count[0] +1
            if self.embarked[i] == 'C' :
                if self.survived[i] == '1' :
                    self.embarked_place_count_survived[1] = self.embarked_place_count_survived[1] +1
                self.embarked_place_count[1] = self.embarked_place_count[1] +1
            if self.embarked[i] == 'Q' :
                if self.survived[i] == '1' :
                    self.embarked_place_count_survived[2] = self.embarked_place_count_survived[2] +1
                self.embarked_place_count[2] = self.embarked_place_count[2] +1

            '''write this code after loop i completes
            self.embarked_influence[0] = self.embarked_place_count_survived[0] / self.embarked_place_count[0] # S
            self.embarked_influence[1] = self.embarked_place_count_survived[1] / self.embarked_place_count[1] # C
            self.embarked_influence[2] = self.embarked_place_count_survived[2] / self.embarked_place_count[2] # Q
            '''


        
        self.male_class_influence[0] = self.male_class[0] / self.male_class_count[0]  
        self.male_class_influence[1] = self.male_class[1] / self.male_class_count[1] 
        self.male_class_influence[2] = self.male_class[2] / self.male_class_count[2]  

        self.female_class_influence[0] = self.female_class[0] / self.female_class_count[0]  
        self.female_class_influence[1] = self.female_class[1] / self.female_class_count[1]  
        self.female_class_influence[2] = self.female_class[2] / self.female_class_count[2]

        self.cabin_alpha_influence[0] = self.cabin_count_alpha_survived[0] / self.cabin_count_alpha[0]
        self.cabin_alpha_influence[1] = self.cabin_count_alpha_survived[1] / self.cabin_count_alpha[1]
        self.cabin_alpha_influence[2] = self.cabin_count_alpha_survived[2] / self.cabin_count_alpha[2]
        self.cabin_alpha_influence[3] = self.cabin_count_alpha_survived[3] / self.cabin_count_alpha[3]
        self.cabin_alpha_influence[4] = self.cabin_count_alpha_survived[4] / self.cabin_count_alpha[4]
        self.cabin_alpha_influence[5] = self.cabin_count_alpha_survived[5] / self.cabin_count_alpha[5]

        self.embarked_influence[0] = self.embarked_place_count_survived[0] / self.embarked_place_count[0]  
        self.embarked_influence[1] = self.embarked_place_count_survived[1] / self.embarked_place_count[1]  
        self.embarked_influence[2] = self.embarked_place_count_survived[2] / self.embarked_place_count[2]  
        self.saving_data_to_excel()

    def saving_data_to_excel(self):
       
        self.ws.cell(row=1, column=1, value='male_class_influence')
        self.ws.cell(row=1, column=2, value='female_class_influence')
        self.ws.cell(row=1, column=3, value='cabin_alpha_influence')
        self.ws.cell(row=1, column=4, value='embarked_influence')

        for i in range(1,4):
            self.ws.cell(row=i+1, column=1, value=self.male_class_influence[i-1])
            self.ws.cell(row=i+1, column=2, value=self.female_class_influence[i-1])
            self.ws.cell(row=i+1, column=4, value=self.embarked_influence[i-1])


        for i in range(2,8):
            self.ws.cell(row=i, column=3, value=self.cabin_alpha_influence[i-2])

        self.wb.save('./datasets/savedProcessed/' + self.fname+'Processed.xlsm') 
        self.file_close()

    def file_close(self):
        self.wb.close()
        print("Record Analysised..!\nOutput saved at " + "./datasets/savedProcessed/"+self.fname+'.xlsm')
        

